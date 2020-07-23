import os
from time import sleep
from random import *

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import ElementNotInteractableException
from bs4 import BeautifulSoup

from openpyxl import load_workbook
from openpyxl import Workbook

# termianl UTF-8 encoding : chcp 65001

##########################################################################
##################### variable related selenium ##########################
##########################################################################
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('lang=ko_KR')
chromedriver_path = "chromedriver"
driver = webdriver.Chrome(os.path.join(os.getcwd(), chromedriver_path), options=options)  # chromedriver 열기
MAX_RAND_INT = 4

##########################################################################
###################### variable related excel ############################
##########################################################################

rowNum = 2
menuRow = 2
save_menuFile_name = "menuLists.xlsx"
menu_wb = load_workbook(os.path.join(os.getcwd(), save_menuFile_name))
menu_ws = menu_wb['Sheet1']

def main():
    global driver, menu_wb

    driver.implicitly_wait(4)  # 렌더링 될때까지 기다린다 4초
    driver.get('https://map.kakao.com/')  # 주소 가져오기

    search("예츠페리")

    driver.quit()
    print("finish")


def search(place):
    global driver

    searchArea = driver.find_element_by_xpath('//*[@id="search.keyword.query"]')
    searchArea.send_keys("스타벅스 서울역동자동점")  # 검색어 입력
    driver.find_element_by_xpath('//*[@id="search.keyword.submit"]').send_keys(Keys.ENTER)  # Enter로 검색
    sleep(1)

    # 검색된 정보가 있는 경우에만 탐색
    # 1번 페이지 cafe list 읽기
    html = driver.page_source

    soup = BeautifulSoup(html, 'html.parser')
    cafeLists = soup.select('.placelist > .PlaceItem')

    crawling(cafeLists)
    menu_wb.save(os.path.join(os.getcwd(), save_menuFile_name))

def crawling(cafeLists):
    for i, cafe in enumerate(cafeLists):
        cafeInfo = []
        cafeInfo.append(cafe.select('.head_item > .tit_name > .link_name')[0].text)  # cafeName
        cafeInfo.append(cafe.select('.info_item > .addr > p')[0].text)  # cafe address
        tempOper = cafe.select('.info_item > .openhour > p > a')  # cafe operation time
        tempPhoneNumber = cafe.select('.info_item > .contact > .phone')  # cafe phone number

        # 운영시간 정보가 없는 경우
        if len(tempOper) != 0:
            cafeInfo.append(tempOper[0].text.strip().replace("\n", ""))
        else:
            cafeInfo.append("None")

        # 전화번호 정보가 없는 경우
        if len(tempPhoneNumber) != 0:
            cafeInfo.append(tempPhoneNumber[0].text)
        else:
            cafeInfo.append("None")

        menuInfos = getMenuInfo(i, driver)
        print(menuInfos)
        saveMenu(menuInfos)

def getMenuInfo(i, driver):
    # 상세페이지로 가서 메뉴찾기
    detail_page_xpath = '//*[@id="info.search.place.list"]/li[' + str(i + 1) + ']/div[5]/div[4]/a[1]'
    driver.find_element_by_xpath(detail_page_xpath).send_keys(Keys.ENTER)
    driver.switch_to.window(driver.window_handles[-1])  # 상세정보 탭으로 변환
    sleep(1)

    menuInfos = []
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')

    menuonlyType = soup.select('.cont_menu > .list_menu > .menuonly_type')
    nophotoType = soup.select('.cont_menu > .list_menu > .nophoto_type')
    photoType = soup.select('.cont_menu > .list_menu > .photo_type')

    if len(menuonlyType) != 0:
        for menu in menuonlyType:
            menuInfos.append(_getMenuInfo(menu))
    elif len(nophotoType) != 0:
        for menu in nophotoType:
            menuInfos.append(_getMenuInfo(menu))
    else:
        for menu in photoType:
            menuInfos.append(_getMenuInfo(menu))

    driver.close()
    driver.switch_to.window(driver.window_handles[0])  # 검색 탭으로 전환

    return menuInfos

def _getMenuInfo(menu):
    menuName = menu.select('.info_menu > .loss_word')[0].text
    menuPrices = menu.select('.info_menu > .price_menu')
    menuPrice = ''

    if len(menuPrices) != 0:
        menuPrice =  menuPrices[0].text.split(' ')[1]

    return [menuName, menuPrice]

def saveMenu(menuInfos):
    global rowNum, menu_wb, menuRow

    # 카페의 모든 메뉴 저장
    for i, menu in enumerate(menuInfos):
        menu_ws.cell(menuRow, 1, rowNum)
        menu_ws.cell(menuRow, 2, i+1)
        for j, info in enumerate(menu):
            menu_ws.cell(menuRow, j+3, info)  # cell(row, col, data)
        menuRow += 1
    menu_wb.save(os.path.join(os.getcwd(), save_menuFile_name))

if __name__ == "__main__":
    main()